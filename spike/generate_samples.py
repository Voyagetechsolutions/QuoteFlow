"""
Generate deliberately MESSY freight rate sheets to test extraction against.

Real carrier/agent rate sheets are not clean tables. This generator reproduces
the noise the extractor must survive:
  - a title block and free-text validity line (no column for it)
  - section headers that ARE the charge-type context ("SEA FREIGHT - FCL")
  - a header row that isn't row 1
  - blank spacer rows
  - mixed currency formats: "USD 1,850.00", "$1850", "1 850"
  - footnote / terms rows that look like data but aren't

Outputs to ../samples/ (gitignored). The generator is committed; the samples
are throwaway and reproducible by running this.

Usage:  python spike/generate_samples.py
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

HERE = os.path.dirname(os.path.abspath(__file__))
SAMPLES = os.path.join(HERE, "..", "samples")

CARRIER = "Kalahari Lines & Forwarding (Pty) Ltd"
VALIDITY = "Rate Validity: 01 Jun 2026 - 31 Aug 2026"

# (section, origin, destination, unit, rate_as_written, currency, remark)
ROWS = [
    ("SEA FREIGHT - FCL", "Durban", "Harare", "per 20ft", "USD 1,850.00", "USD", "Door/Door"),
    ("SEA FREIGHT - FCL", "Durban", "Harare", "per 40ft", "$3,400", "USD", "Door/Door"),
    ("SEA FREIGHT - FCL", "Durban", "Lusaka", "per 20ft", "USD 2,100", "USD", ""),
    ("SEA FREIGHT - FCL", "Cape Town", "Lilongwe", "per 40ft", "4 050", "USD", "Transit 18-21d"),
    ("ROAD FREIGHT - FTL", "Johannesburg", "Harare", "per truck", "ZAR 28,500", "ZAR", "Tri-axle"),
    ("ROAD FREIGHT - FTL", "Johannesburg", "Gaborone", "per truck", "ZAR 16,750.00", "ZAR", ""),
    ("LOCAL CHARGES", "Harare", "", "per shipment", "USD 145", "USD", "Clearing fee"),
    ("LOCAL CHARGES", "Harare", "", "per BL", "USD 65.00", "USD", "Documentation"),
    ("LOCAL CHARGES", "Lusaka", "", "per shipment", "USD 160", "USD", "Clearing fee"),
]

FOOTNOTES = [
    "Rates exclude duties, taxes and statutory charges.",
    "Subject to space and equipment availability at time of booking.",
    "USD rates payable at prevailing exchange rate on date of invoice.",
]


def _ensure_samples_dir() -> None:
    os.makedirs(SAMPLES, exist_ok=True)


def generate_xlsx() -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"

    bold = Font(bold=True)
    section_fill = PatternFill("solid", fgColor="DDDDDD")

    # Title block (not a table) -------------------------------------------
    ws["A1"] = CARRIER
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = VALIDITY
    ws["A2"].font = Font(italic=True)
    # row 3 intentionally blank

    headers = ["Origin", "Destination", "Unit", "Rate", "Currency", "Remarks"]

    r = 4
    current_section = None
    for section, origin, dest, unit, rate, currency, remark in ROWS:
        if section != current_section:
            # blank spacer + section header band
            r += 1
            ws.cell(row=r, column=1, value=section).font = bold
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = section_fill
            r += 1
            # column header row (repeats per section, like real sheets)
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=c, value=h)
                cell.font = bold
            r += 1
            current_section = section
        ws.cell(row=r, column=1, value=origin)
        ws.cell(row=r, column=2, value=dest)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=rate)
        ws.cell(row=r, column=5, value=currency)
        ws.cell(row=r, column=6, value=remark)
        r += 1

    # footnotes ------------------------------------------------------------
    r += 1
    for note in FOOTNOTES:
        ws.cell(row=r, column=1, value=note).font = Font(italic=True, size=9)
        r += 1

    for col, width in zip("ABCDEF", (16, 16, 14, 14, 10, 22)):
        ws.column_dimensions[col].width = width

    out = os.path.join(SAMPLES, "sample_rate_sheet.xlsx")
    wb.save(out)
    return out


# Rows that SHOULD trip review flags, to prove the human-in-the-loop path.
# (section, origin, destination, unit, rate_as_written, currency, remark)
MESSY_ROWS = [
    ("SEA FREIGHT - FCL", "Durban", "Beira", "per 20ft", "POA", "", "On request"),
    ("ROAD FREIGHT - FTL", "", "Ndola", "per truck", "ZAR 22,000", "ZAR", "origin blank"),
    ("SEA FREIGHT - FCL", "Walvis Bay", "Lusaka", "per 20ft", "1850", "", "no currency anywhere"),
]


def generate_messy_xlsx() -> str:
    """A nastier sheet with broken cells AND no validity line -> flags should fire."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"
    bold = Font(bold=True)

    ws["A1"] = CARRIER  # note: deliberately NO validity line
    headers = ["Origin", "Destination", "Unit", "Rate", "Currency", "Remarks"]

    r = 3
    current_section = None
    for section, origin, dest, unit, rate, currency, remark in MESSY_ROWS:
        if section != current_section:
            ws.cell(row=r, column=1, value=section).font = bold
            r += 1
            for c, h in enumerate(headers, start=1):
                ws.cell(row=r, column=c, value=h).font = bold
            r += 1
            current_section = section
        for c, val in enumerate((origin, dest, unit, rate, currency, remark), start=1):
            ws.cell(row=r, column=c, value=val)
        r += 1

    out = os.path.join(SAMPLES, "sample_rate_sheet_messy.xlsx")
    wb.save(out)
    return out


def generate_pdf() -> str:
    out = os.path.join(SAMPLES, "sample_rate_sheet.pdf")
    doc = SimpleDocTemplate(out, pagesize=A4,
                            leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    headers = ["Origin", "Destination", "Unit", "Rate", "Currency", "Remarks"]

    flow = [
        Paragraph(CARRIER, styles["Title"]),
        Paragraph(VALIDITY, styles["Italic"]),
        Spacer(1, 8 * mm),
    ]

    # Group rows into one ruled table per section.
    by_section: dict[str, list[list[str]]] = {}
    order: list[str] = []
    for section, origin, dest, unit, rate, currency, remark in ROWS:
        if section not in by_section:
            by_section[section] = [headers]
            order.append(section)
        by_section[section].append([origin, dest, unit, rate, currency, remark])

    for section in order:
        flow.append(Spacer(1, 4 * mm))
        flow.append(Paragraph(f"<b>{section}</b>", styles["Heading3"]))
        tbl = Table(by_section[section], hAlign="LEFT")
        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DDDDDD")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        flow.append(tbl)

    flow.append(Spacer(1, 8 * mm))
    for note in FOOTNOTES:
        flow.append(Paragraph(f"<i>{note}</i>", styles["Normal"]))

    doc.build(flow)
    return out


def main() -> None:
    _ensure_samples_dir()
    rel = lambda p: os.path.relpath(p, os.path.join(HERE, ".."))
    print("wrote", rel(generate_xlsx()))
    print("wrote", rel(generate_pdf()))
    print("wrote", rel(generate_messy_xlsx()))


if __name__ == "__main__":
    main()
